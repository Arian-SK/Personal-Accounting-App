from PyQt6.QtWidgets import * #to import every tools from QtWidgets
from PyQt6.QtCore import QUrl, Qt, QStringListModel #to define path
from PyQt6.QtGui import QIcon, QKeySequence, QDesktopServices, QPixmap #Icon, shortcut keys, link directions, Images
from PyQt6 import uic # for loading ui seperate from source code(qt designer)
from PyQt6.QtMultimedia import QSoundEffect #for soundtracks
import sys
import re #regex
import time #cooldown
from datetime import datetime, timedelta
import os #to fine the path
import shutil #delte account
import pandas as pd #reading and writing excel
from openpyxl import load_workbook #for searching
import random # for random security question


#finding path to project directory:
my_path = os.path.abspath(os.path.dirname(__file__))

#changing the path into readable form:
project_path = my_path.replace('\\','//')

#path to the background image:
star_theme_path = "background-image : url(" + str(project_path) + "//resources//star theme" + "); background-attachment: fixed"
light_theme_path = "background-image : url(" + str(project_path) + "//resources//light theme" + "); background-attachment: fixed"
dark_theme_path = "background-image : url(" + str(project_path) + "//resources//dark theme" + "); background-attachment: fixed"

class Sound:
    def __init__(self, name = ''):
        self.soundtracks_list = ['background music', 'Alert', 'Correct']
        self.soundtrack = QSoundEffect()
        if name == '':
            pass
        else:
            self.setSoundtrack(name)
    def setSoundtrack(self, name):
        if name in self.soundtracks_list:
            file_path = project_path + '//resources//' + name + '.wav'
            self.soundtrack.setSource(QUrl.fromLocalFile(file_path))
        else:
            print('soundtrack not found')
    def Play(self):
        self.soundtrack.play()
    
    def Mute(self, condition): 
        if condition == False:
            self.soundtrack.setMuted(False)
        elif condition == True:
            self.soundtrack.setMuted(True)
        else:
            raise ValueError('condition must be bool')

    def IsMuted(self):
        return self.soundtrack.isMuted()

    def SetLoop(self, count = 1):
        self.soundtrack.setLoopCount(count)

#main menu ui:
class MainApp(QMainWindow):
    def __init__(self,username = ''):
        super().__init__()

        #ui/tile/icon setup:
        uic.loadUi(project_path + "//MainWindow.ui", self)
        self.setWindowTitle('Personal accountant')
        self.setWindowIcon(QIcon(project_path + "//resources//main icon.ico"))

        #username:
        self.lineUsername.setText(username)
        
        #background image setup:
        self.labelPic = QLabel(self)
        self.labelPic.resize(16777215, 16777215)
        self.labelPic.setStyleSheet(star_theme_path)
        self.labelPic.lower()

        #View list Categories:
        self.model = QStringListModel()
        self.listViewCategories.setModel(self.model)
        self.listViewCategories.setEditTriggers(QListView.EditTrigger.NoEditTriggers)

        #View list Reports:
        self.model_2 = QStringListModel()
        self.listViewReports.setModel(self.model_2)
        self.listViewReports.setEditTriggers(QListView.EditTrigger.NoEditTriggers)

        #View list Search:
        self.model_3 = QStringListModel()
        self.listViewSearch.setModel(self.model_3)
        self.listViewSearch.setEditTriggers(QListView.EditTrigger.NoEditTriggers)


        #others:
        self.buttonMute.setIcon(QIcon(project_path + "//resources//sound.ico"))
        self.buttonInstagram.setIcon(QIcon(project_path + "//resources//Instagram Icon.ico"))
        self.buttonTelegram.setIcon(QIcon(project_path + "//resources//Telegram Icon.ico"))
        self.buttonTwitter.setIcon(QIcon(project_path + "//resources//Twitter Icon.ico"))
        self.comboProfiles.addItem(QIcon(project_path + '//resources//p1.ico'), 'profile1')
        self.comboProfiles.addItem(QIcon(project_path + '//resources//p2.ico'), 'profile2')
        self.comboProfiles.addItem(QIcon(project_path + '//resources//p3.ico'), 'profile3')
        self.comboProfiles.addItem(QIcon(project_path + '//resources//p4.ico'), 'profile4')
        self.comboProfiles.addItem(QIcon(project_path + '//resources//p5.ico'), 'profile5')
        self.light_theme_on = False
        self.dark_theme_on = False
        self.star_theme_on = True

        # set initial volume value
        self.sliderVolume.setValue(50)
        self.update_volume(50)

        #signal handling:

            #Main Menu:
        self.buttonRegCost.clicked.connect(self.go_to_CostsTab)
        self.buttonRegIncome.clicked.connect(self.go_to_IncomeTab)
        self.buttonSettings.clicked.connect(self.go_to_SettingsTab)
        self.buttonSearch.clicked.connect(self.go_to_SearchTab)
        self.buttonReports.clicked.connect(self.go_to_ReportsTab)
        self.buttonCategories.clicked.connect(self.go_to_CategoriesTab)
        self.buttonProfile.clicked.connect(self.go_to_ProfileTab)

            #Profile:
        self.comboProfiles.currentIndexChanged.connect(self.set_profile_pic)
        self.load_user_profile()
        self.set_profile_pic()
        self.buttonChange.clicked.connect(self.change_profile_details)
        self.lineEmail.setEnabled(False)
        self.linePassword.setEnabled(False)
        self.lineFname.setEnabled(False)
        self.lineLname.setEnabled(False)
        self.linePnumber.setEnabled(False)
        self.buttonDelete.clicked.connect(self.delete_account)
        self.buttonDeleteAllSubs.clicked.connect(self.delete_subs)
        self.buttonLogOut.clicked.connect(self.log_out)
        
            #Income:
        self.buttonIncomeSubmit.clicked.connect(self.check_Income_inputs)
        self.labelExceptionIncome.setVisible(False)

            #Categories:
        self.buttonCategorySubmit.clicked.connect(self.addCategory)
        self.update_list_view_category()

            #back buttons:
        self.buttonBackFromIncome.clicked.connect(self.go_to_MainMenu)
        self.buttonBackFromCategories.clicked.connect(self.go_to_MainMenu)
        self.buttonBackFromProfile.clicked.connect(self.go_to_MainMenu)
        self.buttonBackFromSettings.clicked.connect(self.go_to_MainMenu)
        self.buttonBackFromCosts.clicked.connect(self.go_to_MainMenu)
        self.buttonBackFromSearch.clicked.connect(self.go_to_MainMenu)
        self.buttonBackFromReports.clicked.connect(self.go_to_MainMenu)

            #Cost:
        self.buttonCostSubmit.clicked.connect(self.check_Cost_inputs)
        self.labelExceptionCost.setVisible(False)

            #Report:
        self.load_excel_Incomes()
        self.load_excel_Costs()
        self.buttonReportsSubmit.clicked.connect(self.perform_reports)
        self.buttonGroupReports = QButtonGroup()
        self.buttonGroupReports.addButton(self.radioReportsPastD)
        self.buttonGroupReports.addButton(self.radioReportsPastM)
        self.buttonGroupReports.addButton(self.radioReportsPastY)
        self.buttonGroupReports.addButton(self.radioReportsNone)
        self.radioReportsNone.setChecked(True)
        self.buttonReportsRange.clicked.connect(self.get_integer_values2)
        self.buttonReportsReset.clicked.connect(self.reset_reports)
        self.first_range1 = 0
        self.second_range2 = 0

            #Search:
        self.buttonSearchSubmit.clicked.connect(self.begin_search)
        self.buttonGroupSearch1 = QButtonGroup()
        self.buttonGroupSearch1.addButton(self.radioSearchIncomesOnly)
        self.buttonGroupSearch1.addButton(self.radioSearchCostsOnly)
        self.buttonGroupSearch1.addButton(self.radioSearchNone1)
        self.radioSearchNone1.setChecked(True)
        self.buttonGroupSearch2 = QButtonGroup()
        self.buttonGroupSearch2.addButton(self.radioSearchPast1)
        self.buttonGroupSearch2.addButton(self.radioSearchPast3)
        self.buttonGroupSearch2.addButton(self.radioSearchNone2)
        self.radioSearchNone2.setChecked(True)
        self.buttonSearchRange.clicked.connect(self.get_integer_values)
        self.buttonSearchReset.clicked.connect(self.reset_search)
        self.firt_range = 0
        self.second_range = 0

            #setting:
        self.buttonMute.clicked.connect(windowLogin.play_mute_background)
        self.sliderVolume.valueChanged.connect(self.update_volume)
        self.buttonChangeTheme.clicked.connect(self.change_theme)
        self.buttonAbout.clicked.connect(self.show_message_about)
        self.buttonDonate.clicked.connect(self.open_link_donation)
        self.buttonInstagram.clicked.connect(self.open_link_instagram)
        self.buttonTelegram.clicked.connect(self.open_link_telegram)
        self.buttonTwitter.clicked.connect(self.open_link_twitter)
        self.tabWidget.tabBar().hide()
        
        #exception handling:
        self.labelExceptionCategory.setVisible(False)
        self.labelExceptionProfile.setVisible(False)

    #reinitialize the class
    def reinit(self):
        self.__init__()

    #profile tab:
    def load_user_profile(self):
            self.username = windowLogin.username
            df = pd.read_excel(project_path + '//database//members_info.xlsx')
            user_row = df[(df['username'] == self.username)]
            email = str(user_row['email']).split()
            password = str(user_row['password']).split()
            fname = str(user_row['first name']).split()
            lname = str(user_row['last name']).split()
            pnumber = str(user_row['phone number']).split()
            self.email = email[1]
            self.password = password[1]
            self.fname = fname[1]
            self.lname = lname[1]
            self.pnumber = '0' + pnumber[1]
            self.lineEmail.setText(self.email)
            self.linePassword.setText(self.password)
            self.lineUsername.setText(self.username)
            self.lineFname.setText(self.fname)
            self.lineLname.setText(self.lname)
            self.linePnumber.setText(self.pnumber)

    def set_profile_pic(self):
        profile_name = self.comboProfiles.currentText()
        pixmap = QPixmap(project_path + '//resources//' + profile_name + '.jpg')
        scaled_pixmap = pixmap.scaled(self.labelProfilePic.size(), Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        self.labelProfilePic.setPixmap(scaled_pixmap)
        self.labelProfilePic.setAlignment(Qt.AlignmentFlag.AlignCenter)
    
    def change_profile_details(self):
        if self.buttonChange.text() == 'Change':
            #change
            self.lineEmail.setEnabled(True)
            self.linePassword.setEnabled(True)
            self.lineFname.setEnabled(True)
            self.lineLname.setEnabled(True)
            self.linePnumber.setEnabled(True)
            self.buttonChange.setText('Submit')
        else:
            #submit
            self.lineEmail.setEnabled(False)
            self.linePassword.setEnabled(False)
            self.lineFname.setEnabled(False)
            self.lineLname.setEnabled(False)
            self.linePnumber.setEnabled(False)
            self.buttonChange.setText('Change')
            self.check_changed_details()

    def check_changed_details(self):
        new_email = self.lineEmail.text()
        new_password = self.linePassword.text()
        new_fname = self.lineFname.text()
        new_lname = self.lineLname.text()
        new_pnumber = self.linePnumber.text()
        valid_email = r'^[a-zA-Z0-9._%+-]+@(gmail|yahoo)\.com$'
        valid_password = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$'
        valid_pnumber = r'^09\d{9}$'
        valid_fname = r'^[a-zA-Z]+$'
        valid_lname = r'^[a-zA-Z]+$'
        df = pd.read_excel(project_path + '//database//members_info.xlsx')
        self.labelExceptionProfile.setText('')
        if new_email != self.email:
            if re.match(valid_email, new_email):
                if new_email in df['email'].values:
                    self.labelExceptionProfile.setVisible(True)
                    self.labelExceptionProfile.setText('email already in use')
                else:
                    self.labelExceptionProfile.setVisible(False)
                    self.labelExceptionProfile.setText('')
                    self.save_changed_email(new_email)
        
        if new_password != self.password:
            if re.match(valid_password, new_password):
                self.labelExceptionProfile.setVisible(False)
                self.labelExceptionProfile.setText('')
                self.save_changed_password(new_password)

            else:
                self.labelExceptionProfile.setVisible(True)
                self.labelExceptionProfile.setText(str(self.labelExceptionProfile.text()) + ' invalid password')

        if new_fname != self.fname:
            if re.match(valid_fname, new_fname):
                self.labelExceptionProfile.setVisible(False)
                self.labelExceptionProfile.setText('')
                self.save_changed_fname(new_fname)
            else:
                self.labelExceptionProfile.setVisible(True)
                self.labelExceptionProfile.setText(str(self.labelExceptionProfile.text()) + ' invalid fname')
        
        if new_lname != self.lname:
            if re.match(valid_lname, new_lname):
                self.labelExceptionProfile.setVisible(False)
                self.labelExceptionProfile.setText('')
                self.save_changed_lname(new_lname)
            else:
                self.labelExceptionProfile.setVisible(True)
                self.labelExceptionProfile.setText(str(self.labelExceptionProfile.text()) + ' invalid lname')

        if new_pnumber != self.pnumber:
            if re.match(valid_pnumber, new_pnumber):
                if new_pnumber in df['phone number'].values:
                    self.labelExceptionProfile.setVisible(True)
                    self.labelExceptionProfile.setText(str(self.labelExceptionProfile.text()) + ' phone number already in use')
                else:
                    self.labelExceptionProfile.setVisible(False)
                    self.labelExceptionProfile.setText('')
                    self.save_changed_pnumber(new_email)

        self.load_user_profile()

    def save_changed_email(self, new_email):
        file_path = project_path + '//database//members_info.xlsx'
        df = pd.read_excel(project_path + '//database//members_info.xlsx')
        df.loc[df['username'] == self.username, 'email'] = new_email
        df.to_excel(file_path, index=False)
    
    def save_changed_password(self, new_password):
        file_path = project_path + '//database//members_info.xlsx'
        df = pd.read_excel(project_path + '//database//members_info.xlsx')
        df.loc[df['username'] == self.username, 'password'] = new_password
        df.to_excel(file_path, index=False)
    
    def save_changed_fname(self, new_fname):
        file_path = project_path + '//database//members_info.xlsx'
        df = pd.read_excel(project_path + '//database//members_info.xlsx')
        df.loc[df['username'] == self.username, 'first name'] = new_fname
        df.to_excel(file_path, index=False)

    def save_changed_lname(self, new_lname):
        file_path = project_path + '//database//members_info.xlsx'
        df = pd.read_excel(project_path + '//database//members_info.xlsx')
        df.loc[df['username'] == self.username, 'last name'] = new_lname
        df.to_excel(file_path, index=False)

    def save_changed_pnumber(self, new_pnumber):
        file_path = project_path + '//database//members_info.xlsx'
        df = pd.read_excel(project_path + '//database//members_info.xlsx')
        df.loc[df['username'] == self.username, 'phone number'] = new_pnumber
        df.to_excel(file_path, index=False)

    def delete_account(self):
        reply = QMessageBox.question(
            self, "Delete Account", f"Are you sure you want to delete this account ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            file_path = project_path + '//database//members_info.xlsx'
            user_folder_path = project_path + '//database//reports//' + self.username
            df = pd.read_excel(file_path)
            username_column = 'username'
            user_to_delete = self.username
            df_filtered = df[df[username_column] != user_to_delete]
            df_filtered.to_excel(file_path, index=False)
            if os.path.isdir(user_folder_path):
                shutil.rmtree(user_folder_path)
            windowMain.close()
            windowLogin.show()
        
    def delete_subs(self):
        reply = QMessageBox.question(
        self, "Delete submissions", f"Are you sure you want to delete all the submissions ?",
        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.user_folder_path = project_path + '//database//reports//' + self.username
            if os.path.isdir(self.user_folder_path):
                shutil.rmtree(self.user_folder_path)
                os_directory_path = self.user_folder_path
                os.makedirs(os_directory_path, exist_ok=True)
                self.make_incomes_excel()
                self.make_costs_excel()
                self.make_categories_excel()
                self.reset_combo_source()
                self.update_list_view_reports([],0)
                self.update_list_view_category(0)

    def make_incomes_excel(self):
        df = pd.DataFrame(columns=['Income', 'Date', 'Source', 'Details', 'Type', 'submit date'])
        file_path = self.user_folder_path + '//Incomes.xlsx'
        df.to_excel(file_path, index=False, engine='openpyxl')

    def make_costs_excel(self):
        df = pd.DataFrame(columns=['Cost', 'Date', 'Source', 'Details', 'Type', 'submit date'])
        file_path = self.user_folder_path + '//Costs.xlsx'
        df.to_excel(file_path, index=False, engine='openpyxl')

    def make_categories_excel(self):
        df = pd.DataFrame(columns=['Categories', 'submit date'])
        file_path = self.user_folder_path + '//Categories.xlsx'
        df.to_excel(file_path, index=False, engine='openpyxl')
    
    def reset_combo_source(self):
        default_items = ['Groceries', 'Refueling', 'Decorations', 'Installment']
        self.comboIncomeSource.clear()
        self.comboIncomeSource.addItems(default_items)
        self.comboIncomeSource.setCurrentIndex(0)
        self.comboCostSource.clear()
        self.comboCostSource.addItems(default_items)
        self.comboCostSource.setCurrentIndex(0)
        self.comboReportsSource.clear()
        self.comboReportsSource.addItems(default_items)
        self.comboReportsSource.setCurrentIndex(0)
        
    def log_out(self):
        print('logging out...')
        windowMain.close()
        windowLogin.show()

    #main menu tab:
    def go_to_MainMenu(self):
        self.tabWidget.setCurrentIndex(0)

    def go_to_IncomeTab(self):
        self.tabWidget.setCurrentIndex(1)

    def go_to_CostsTab(self):
        self.tabWidget.setCurrentIndex(2)
    
    def go_to_SearchTab(self):
        self.tabWidget.setCurrentIndex(3)
    
    def go_to_CategoriesTab(self):
        self.tabWidget.setCurrentIndex(4)
    
    def go_to_ReportsTab(self):
        self.tabWidget.setCurrentIndex(5)

    def go_to_SettingsTab(self):
        self.tabWidget.setCurrentIndex(6)
    
    def go_to_ProfileTab(self):
        self.tabWidget.setCurrentIndex(7)

    #income tab:
    def reset_Income_inputs(self):
        self.lineIncome.setText('')
        self.lineIncomeDate.setText('')
        self.lineIncomeDetails.setText('')
        self.lineIncomeSource.setText('')
    
    def check_Income_inputs(self):
        if self.check_Income() == False:
            windowLogin.play_wrong()
            self.labelExceptionIncome.setVisible(True)
            self.labelExceptionIncome.setText('invalid input for income')
            return
        else:
            self.labelExceptionIncome.setVisible(False)
            self.labelExceptionIncome.setText('')
        if self.check_Income_Date() == False:
            windowLogin.play_wrong()
            self.labelExceptionIncome.setVisible(True)
            self.labelExceptionIncome.setText('invalid date')
            return
        else:
            self.labelExceptionIncome.setVisible(False)
            self.labelExceptionIncome.setText('')

        self.submit('income')

    def check_Income(self):
        valid_Income = r'^\d+(\.\d+)?$'
        if re.match(valid_Income, self.lineIncome.text()):   
            return True
        else:
            return False

    def check_Income_Date(self):
        valid_Income_Date = r'^\d{4}/\d{2}/\d{2}$'  
        if re.match(valid_Income_Date, self.lineIncomeDate.text()):
            year, month, day = self.lineIncomeDate.text().split('/')
            if 0 < int(day) <= 31 and 0 < int(month) <= 12:
                return True
            else:
                return False
        else:
            return False

    #Cost tab:
    def check_Cost_inputs(self):
        if self.check_Cost() == False:
            windowLogin.play_wrong()
            self.labelExceptionCost.setVisible(True)
            self.labelExceptionCost.setText('invalid input for cost')    
            return
        else:
            self.labelExceptionCost.setVisible(False)
            self.labelExceptionCost.setText('') 
        if self.check_Cost_Date() == False:
            windowLogin.play_wrong()
            self.labelExceptionCost.setVisible(True)
            self.labelExceptionCost.setText('invalid date') 
            return
        else:
            self.labelExceptionCost.setVisible(False)
            self.labelExceptionCost.setText('') 

        self.submit('cost')

    def check_Cost(self):
        valid_Cost = r'^\d+(\.\d+)?$'
        if re.match(valid_Cost, self.lineCost.text()):
            return True
        else:
            return False

    def check_Cost_Date(self):
        valid_Cost_Date = r'^\d{4}/\d{2}/\d{2}$'
        if re.match(valid_Cost_Date, self.lineCostDate.text()):
            year, month, day = self.lineCostDate.text().split('/')
            if 0 < int(day) <= 31 and 0 < int(month) <= 12:
                return True
            else:
                return False
        else:
            return False

    #Submiting:
    def submit(self, Type):
        current_time = time.localtime()
        self.current_time = time.strftime('%Y/%m/%d', current_time)
        if Type == 'income':
            self.Income = self.lineIncome.text()
            self.IncomeDate = self.lineIncomeDate.text()
            self.IncomeSource = self.comboIncomeSource.currentText()
            self.IncomeDetails = self.lineIncomeDetails.text()
            self.IncomeType = self.comboIncomeType.currentText()
            
            info = {
            "Income": [self.Income],
            "Date": [self.IncomeDate],
            "Source": [self.IncomeSource],
            "Details": [self.IncomeDetails],
            "Type": [self.IncomeType],
            "submit date": [self.current_time]
            }
            
            new_data = pd.DataFrame(info)

            file_path = project_path + "//database//reports//" + self.username + "//incomes.xlsx"

        elif Type == 'cost':
            self.Cost = self.lineCost.text()
            self.CostDate = self.lineCostDate.text()
            self.CostSource = self.comboCostSource.currentText()
            self.CostDetails = self.lineCostDetails.text()
            self.CostType = self.comboCostType.currentText()

            info = {
            "Cost": [self.Cost],
            "Date": [self.CostDate],
            "Source": [self.CostSource],
            "Details": [self.CostDetails],
            "Type": [self.CostType],
            "submit date": [self.current_time]
            }

            new_data = pd.DataFrame(info)

            file_path = project_path + "//database//reports//" + self.username + "//costs.xlsx"

        elif Type == 'category':
            category = self.lineNewCategory.text()
            self.new_category = category.capitalize()

            info = {
                "Categories": [self.new_category],
                "submit date": [self.current_time]
            }

            new_data = pd.DataFrame(info)

            file_path = project_path + "//database//reports//" + self.username + "//categories.xlsx"

        else:
            self.show_message_unsuccessful()
        
        if os.path.exists(file_path):
            existing_data = pd.read_excel(file_path)
            updated_data = pd.concat([existing_data, new_data], ignore_index=True)

        else:
            updated_data = new_data
        
        updated_data.to_excel(file_path, index=False)
        mod_info = str(info).replace('[','').replace(']','').replace("'","").replace('}','').replace('{','').replace(', ','\n')
        self.update_list_view_reports(mod_info)
        self.show_message_successful()
        self.reset_inputs()

    def show_message_successful(self):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setWindowTitle('Success')
        msg_box.setText('submission was successful')
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.exec()

    def show_message_unsuccessful(self):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Warning)
        msg_box.setWindowTitle('Error')
        msg_box.setText('submission was unsuccessful')
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.exec()

    def reset_inputs(self):
            self.lineIncome.setText('')
            self.lineIncomeDate.setText('')
            self.lineIncomeDetails.setText('')
            self.lineCost.setText('')
            self.lineCostDate.setText('')
            self.lineCostDetails.setText('')

    #search tab:
    def begin_search(self):
        self.perform_search()

    def reset_search(self):
        self.radioSearchNone1.setChecked(True)
        self.radioSearchNone2.setChecked(True)
        self.first_range = 0
        self.second_range = 0
        self.labelSearchRange.setText('Range')

    def get_integer_values(self):
        # First dialog
        first_value, ok1 = QInputDialog.getInt(None, "Input Dialog", "Enter first integer value:")
        if not ok1:
            QMessageBox.critical(None, "Error", "First input was cancelled.")
            return
        
        while True:
            # Second dialog
            second_value, ok2 = QInputDialog.getInt(None, "Input Dialog", "Enter second integer value:")
            if not ok2:
                QMessageBox.critical(None, "Error", "Second input was cancelled.")
                return
            
            if second_value > first_value:
                break
            else:
                QMessageBox.warning(None, "Input Error", "Second value must be greater than the first value. Please try again.")
        
        self.labelSearchRange.setText('range: ' + str(first_value) + ' - ' + str(second_value))
        self.first_range = first_value
        self.second_range = second_value

    def perform_search(self):
        self.excel_path_income = project_path + '//database//reports//' + self.username + '//incomes.xlsx'
        self.excel_path_cost = project_path + '//database//reports//' + self.username + '//costs.xlsx'
        self.excel_path_categories = project_path + '//database//reports//' + self.username + '//categories.xlsx'

        self.model_3.setStringList([''])
        self.results_combined = []

        #None1,None2
        if self.radioSearchNone1.isChecked() and self.radioSearchNone2.isChecked():
            if os.path.exists(self.excel_path_income):
                self.Search(self.excel_path_income)
            if os.path.exists(self.excel_path_cost):
                self.Search(self.excel_path_cost)
        
        #None1,Past 24 hours
        if self.radioSearchNone1.isChecked() and self.radioSearchPast1.isChecked():
            if os.path.exists(self.excel_path_income):
                self.Search(self.excel_path_income, 1)
            if os.path.exists(self.excel_path_cost):
                self.Search(self.excel_path_cost, 1)

        #None1,Past 3 days
        if self.radioSearchNone1.isChecked() and self.radioSearchPast3.isChecked():
            if os.path.exists(self.excel_path_income):
                self.Search(self.excel_path_income, 3)
            if os.path.exists(self.excel_path_cost):
                self.Search(self.excel_path_cost, 3)

        #IncomesOnly,None2
        if self.radioSearchIncomesOnly.isChecked() and self.radioSearchNone2.isChecked():
            if os.path.exists(self.excel_path_income):
                self.Search(self.excel_path_income)

        #IncomesOnly, Past 24 hours
        if self.radioSearchIncomesOnly.isChecked() and self.radioSearchPast1.isChecked():
            if os.path.exists(self.excel_path_income):
                self.Search(self.excel_path_income, 1)
        
        #IncomesOnly, Past 3 days
        if self.radioSearchIncomesOnly.isChecked() and self.radioSearchPast3.isChecked():
            if os.path.exists(self.excel_path_income):
                self.Search(self.excel_path_income, 3)

        #CostsOnly, None2
        if self.radioSearchCostsOnly.isChecked() and self.radioSearchNone2.isChecked():
            if os.path.exists(self.excel_path_cost):
                self.Search(self.excel_path_cost)

        #CostsOnly, Past 24 hours
        if self.radioSearchCostsOnly.isChecked() and self.radioSearchPast1.isChecked():
            if os.path.exists(self.excel_path_cost):
                self.Search(self.excel_path_cost, 1)

        #CostsOnly, Past 3 days
        if self.radioSearchCostsOnly.isChecked() and self.radioSearchPast3.isChecked():
            if os.path.exists(self.excel_path_cost):
                self.Search(self.excel_path_cost, 3)
        
        self.update_list_view_search(self.results_combined)

    def Search(self, file_path, time_range = ''):
        searched_string = self.lineSearch.text()
        result = []
        if searched_string:
            #search the key word:
            result = self.search_in_excel(file_path, searched_string)
            if result:
                self.result_display = []
                new_results = []

                #check if it's withing the past 24 hours:
                if time_range == 1:
                    for i in range(len(result)):
                        current_time_str = result[i][1].strip("'Date: ")
                        if self.is_within_past_days(current_time_str, 1):
                            new_results.append(result[i])
                    result = new_results

                #check if it's withing the past 3 days:
                elif time_range == 3:
                    for i in range(len(result)):
                        current_time_str = result[i][1].strip("'Date: ")
                        if self.is_within_past_days(current_time_str, 3):
                            new_results.append(result[i])
                    result = new_results
                
                #money range check:
                if self.second_range > 0:
                    for i in range(len(result)):
                        money = result[i][0].strip("'Income: ").strip("'Cost: ")
                        if self.is_within_range(money):
                            new_results.append(result[i])
                    result = new_results
                new_results = []

                #if it's empty don't display, if not add \n to every element:
                if len(result) > 0:
                    for i in result:
                        new_results.append(i)
                        new_results.append('--------------------------------------')
                    result = new_results
                    self.result_display.append("\n".join([str(row) for row in result]))
                    self.results_combined.append(self.result_display)

    def is_within_range(self, money_str):
        money = int(money_str)
        return self.first_range < money < self.second_range

    def is_within_past_days(self, date_str, past_days):
        #Convert string to datetime object
        date_obj = datetime.strptime(date_str, "%Y/%m/%d")
        
        #Get current time
        current_time = datetime.now()
        
        # Calculate difference
        difference = current_time - date_obj
        
        # Define timedelta for x days
        Days = timedelta(days = past_days)

        # Check if difference is less than x days
        return difference < Days

    def search_in_excel(self, file_path, searched_string):
        workbook = load_workbook(file_path)
        sheet = workbook.active

        search_results = []
        column_headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(values_only=True):
            if any(searched_string.lower() in str(cell).lower() for cell in row):
                row_data = [f"{column_headers[i]}: {cell}" for i, cell in enumerate(row)]
                search_results.append(row_data)

        return search_results

    def update_list_view_search(self, List):
        string_list = self.model_3.stringList()
        for i in List:
            for j in i:
                string_list.append(str(j).replace("'", "").replace('[','').replace(']','').replace(',','\n'))
        self.model_3.setStringList(string_list)

    #Categories tab:
    def addCategory(self):
        Category = self.lineNewCategory.text()
        valid_category = r'^[a-zA-Z]+$'
        if Category.isalpha() and len(Category) <= 15:
            New_Category = Category.capitalize()
        else:
            self.labelExceptionCategory.setVisible(True)
            self.labelExceptionCategory.setText('invalid Category')
            return
        if re.match(valid_category , New_Category):
            if self.comboIncomeSource.findText(New_Category) == -1:
                self.comboIncomeSource.addItem(New_Category)
                self.comboCostSource.addItem(New_Category)
                self.submit('category')
                self.lineNewCategory.setText('')
                self.update_list_view_category()
                self.labelExceptionCategory.setVisible(False)
                self.labelExceptionCategory.setText('')
                return
            else:
                self.labelExceptionCategory.setVisible(True)
                self.labelExceptionCategory.setText('Category already exists!')
                return
        else:
            self.labelExceptionCategory.setVisible(True)
            self.labelExceptionCategory.setText('only English characters are acceptable')
            return 
    
    def update_list_view_category(self, code = 1):
        #code 0 is for clearing the view list
        if code == 0:
            self.model.setStringList([])
        else:
            items = [self.comboIncomeSource.itemText(i) for i in range(self.comboIncomeSource.count())]
            self.model.setStringList(items)
            self.comboReportsSource.addItems(items)

    #reports tab:
    def reset_reports(self):
        self.radioReportsNone.setChecked(True)
        self.labelReportsRange.setText('')
        self.first_range2 = 0
        self.second_range2 = 0
        self.comboReportsType.setCurrentText('Type(none)')
        self.comboReportsSource.setCurrentText('Source(none)')
        self.labelReportsRange.setText('Range')

    def update_list_view_reports(self, info, code = 1):
        if code == 0:
            self.model_2.setStringList([])
        else:
            string_list = self.model_2.stringList()
            string_list.append(info)
            string_list.append('--------------------------------------')
            self.model_2.setStringList(string_list)

    def load_excel_Incomes(self):
        file_path = project_path + "//database//reports//" + self.username + "//incomes.xlsx"

        if os.path.exists(file_path):
            df = pd.read_excel(file_path)
            incomes_list = df.values.tolist()
            for i in range(len(incomes_list)):
                info = {
                "Income": [incomes_list[i][0]],
                "Date": [incomes_list[i][1]],
                "Source": [incomes_list[i][2]],
                "Details": [incomes_list[i][3]],
                "Type": [incomes_list[i][4]],
                "submit date": [incomes_list[i][5]]
                }
                mod_info = str(info).replace('[','').replace(']','').replace("'","").replace('}','').replace('{','').replace(', ','\n')
                self.update_list_view_reports(mod_info)
        else:
            return

    def load_excel_Costs(self):
        file_path = project_path + "//database//reports//" + self.username + "//costs.xlsx"
        if os.path.exists(file_path):
            df = pd.read_excel(file_path)
            costs_list = df.values.tolist()
            for i in range(len(costs_list)):
                info = {
                "Cost": [costs_list[i][0]],
                "Date": [costs_list[i][1]],
                "Source": [costs_list[i][2]],
                "Details": [costs_list[i][3]],
                "Type": [costs_list[i][4]],
                "submit date": [costs_list[i][5]]
                }
                mod_info = str(info).replace('[','').replace(']','').replace("'","").replace('}','').replace('{','').replace(', ','\n')
                self.update_list_view_reports(mod_info)
        else:
            return

    def get_integer_values2(self):
        first_value, ok1 = QInputDialog.getInt(None, "Input Dialog", "Enter first integer value:")
        if not ok1:
            QMessageBox.critical(None, "Error", "First input was cancelled.")
            return
        
        while True:
            # Second dialog
            second_value, ok2 = QInputDialog.getInt(None, "Input Dialog", "Enter second integer value:")
            if not ok2:
                QMessageBox.critical(None, "Error", "Second input was cancelled.")
                return
            
            if second_value > first_value:
                break
            else:
                QMessageBox.warning(None, "Input Error", "Second value must be greater than the first value. Please try again.")
        
        self.labelReportsRange.setText('range: ' + str(first_value) + ' - ' + str(second_value))
        self.first_range2 = first_value
        self.second_range2 = second_value

    def perform_reports(self):
        self.excel_path_income = project_path + '//database//reports//' + self.username + '//incomes.xlsx'
        self.excel_path_cost = project_path + '//database//reports//' + self.username + '//costs.xlsx'
        self.excel_path_categories = project_path + '//database//reports//' + self.username + '//categories.xlsx'

        self.model_2.setStringList([''])
        self.results_combined = []

        if self.radioReportsNone.isChecked():
            self.filter_reports(self.excel_path_income)
            self.filter_reports(self.excel_path_cost)

        if self.radioReportsPastD.isChecked():
            self.filter_reports(self.excel_path_income,'D')
            self.filter_reports(self.excel_path_cost,'D')

        if self.radioReportsPastM.isChecked():
            self.filter_reports(self.excel_path_income,'M')
            self.filter_reports(self.excel_path_cost,'M')

        if self.radioReportsPastY.isChecked():
            self.filter_reports(self.excel_path_income,'Y')
            self.filter_reports(self.excel_path_cost,'Y')

        self.update_list_view_reports_filtered(self.results_combined)
        
    def filter_reports(self, file_path, time_range = ''):
        result = []
        searched_string = ''
        if self.comboReportsSource.currentText() != 'Source(none)':
            searched_string = self.comboReportsSource.currentText()

        if searched_string:
            result = self.search_in_excel(file_path, searched_string)
        else:
            result = self.read_excel_to_list(file_path)

        if result:
            self.result_display = []
            new_results = []
        
            if time_range == 'D':
                for i in range(len(result)):
                    current_time_str = result[i][1].strip("'Date: ")
                    if self.is_within_past_days(current_time_str, 1):
                        new_results.append(result[i])
                result = new_results
            elif time_range == 'M':
                for i in range(len(result)):
                    current_time_str = result[i][1].strip("'Date: ")
                    if self.is_within_past_days(current_time_str, 30):
                        new_results.append(result[i])
                result = new_results
            elif time_range == 'Y':
                for i in range(len(result)):
                    current_time_str = result[i][1].strip("'Date: ")
                    if self.is_within_past_days(current_time_str, 365):
                        new_results.append(result[i])
                result = new_results
            
            new_results = []
            if self.second_range2 > 0:
                for i in range(len(result)):
                    money = result[i][0].strip("'Income: ").strip("'Cost: ")
                    if self.is_within_range2(money):
                        new_results.append(result[i])
                result = new_results
            
            new_results = []
            if self.comboReportsType.currentText() != 'Type(none)':
                Type = self.comboReportsType.currentText()
                for i in range(len(result)):
                    if Type in result[i][4]:
                        new_results.append(result[i])
                result = new_results
            new_results = []

            if len(result) > 0:
                for i in result:
                    new_results.append(i)
                    new_results.append('--------------------------------------')
                result = new_results
                self.result_display.append("\n".join([str(row) for row in result]))
                self.results_combined.append(self.result_display)

    def read_excel_to_list(self, file_path):
        try:
            # Read the Excel file
            df = pd.read_excel(file_path)
            # Get column names
            columns = list(df.columns)
            # Initialize 2D list with column names
            data = [columns]
            # Iterate over each row
            for _, row in df.iterrows():
                # Create a list for the row
                row_list = []
                # Iterate over each column
                for col in columns:
                    # Construct element with column name and cell value
                    element = f"{col}: {row[col]}"
                    row_list.append(element)
                # Append row list to the 2D list
                data.append(row_list)
            data.pop(0)
            return data
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return None

    def is_within_range2(self, money):
        return self.first_range1 < int(money) < self.second_range2

    def update_list_view_reports_filtered(self, List):
        string_list = self.model_2.stringList()
        for i in List:
            for j in i:
                string_list.append(str(j).replace("'", "").replace('[','').replace(']','').replace(',','\n'))
        self.model_2.setStringList(string_list)

    #settings:
    def change_theme(self):
        if self.star_theme_on:
            self.light_theme_on = True
            self.star_theme_on = False
            path = light_theme_path
            self.labelTheme.setText('Light')

        elif self.light_theme_on:
            self.light_theme_on = False
            self.dark_theme_on = True
            path = dark_theme_path
            self.labelTheme.setText('Dark')

        elif self.dark_theme_on:
            self.star_theme_on = True
            self.dark_theme_on = False
            path = star_theme_path
            self.labelTheme.setText('Stars')
        
        self.labelPic.setStyleSheet(path)
        
    def show_message_about(self):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setWindowTitle('About')
        msg_box.setText('made by Erfan Ghasemian and Arian Saeed Kondori')
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.exec()

    def update_volume(self, value):
        self.labelVolumeLevel.setText("Volume: " + str(value))
        backgroundSound.soundtrack.setVolume(value / 100.0)

    def open_link_donation(self):
        QDesktopServices.openUrl(QUrl('https://developer.paypal.com/'))

    def open_link_instagram(self):
        QDesktopServices.openUrl(QUrl('https://www.instagram.com/'))
    
    def open_link_telegram(self):
        QDesktopServices.openUrl(QUrl('https://telegram.org/'))
    
    def open_link_twitter(self):
        QDesktopServices.openUrl(QUrl('https://twitter.com/'))


#sign up ui:        
class SignUp(QWidget):
    def __init__(self):
        super().__init__()

        #ui/tile/icon setup:
        uic.loadUi(project_path + "//SignUpPage.ui", self)
        self.setWindowIcon(QIcon(project_path + "//resources//signup icon.ico"))
        self.buttonMute.setIcon(QIcon(project_path + "//resources//sound.ico"))
        self.setWindowTitle('Sign Up')

        #background image setup:
        self.labelPic = QLabel(self)
        self.labelPic.resize(16777215, 16777215)
        self.labelPic.setStyleSheet(star_theme_path)
        self.labelPic.lower()
        
        #exception handling label:
        self.labelException.setVisible(False)

        #signal handling:
        self.buttonSignUp.clicked.connect(self.check)
        self.buttonLogin.clicked.connect(self.open_login_page)
        self.buttonMute.clicked.connect(windowLogin.play_mute_background)
        self.lineCity.textChanged.connect(self.change_text)
        self.buttonSignUp.setShortcut("Return")

        #attributes:
        self.city_list = ['Tehran', 'ُSari', 'Karaj', 'Babol', 'Esfahan',
         'Shiraz', 'Yazd', 'Tabriz', 'Kerman', 'Qom', 'Mashhad', 'Ahvaz',
          'Zahedan', 'Kashan', 'Arak', 'Zanjan', 'Ardabil', 'Rasht', 'Amirkola',
           'Hamedan', 'Gorgan', 'Eslamshahr', 'Bandar Abbas', 'Oromieh']
        
        #auto completer:
        self.completer = QCompleter(self.city_list)
        self.lineCity.setCompleter(self.completer)

        #sound initilization:
        self.wrong_sound = Sound('Alert')
        self.wrong_sound_isMuted = False
        self.correct_sound = Sound('Correct')
        self.correct_sound_isMuted = False

        #password line:
        self.linePass.textChanged.connect(self.update_labelPassHint)
        self.linePass.focusInEvent = self.on_password_focus_in
        self.linePass.focusOutEvent = self.on_password_focus_out
        self.labelPassHint.setVisible(False)
        self.buttonToggleEcho.clicked.connect(self.toggle_echo_mode)
        self.buttonToggleEcho.setIcon(QIcon(project_path + "//resources//close eye.ico"))
        
    def change_text(self):
        city = self.lineCity.text()
        new_city = city.capitalize()
        self.lineCity.setText(new_city)

    def open_login_page(self):
        windowLogin.show()
        windowSignUp.close() 

    def play_wrong(self):
        if self.wrong_sound_isMuted == False:
            self.wrong_sound.Play()
    
    def play_correct(self):
        if self.correct_sound_isMuted == False:
            self.correct_sound.Play()

    #function to check inputs:
    def check(self):
        self.labelException.setVisible(True)
        if self.check_fname() == False:
            self.play_wrong()
            return
        if self.check_lname() == False:
            self.play_wrong()
            return
        if self.check_pnumber() == False:
            self.play_wrong()
            return
        if self.check_username() == False:
            self.play_wrong()
            return
        if self.check_email() == False:
            self.play_wrong()
            return
        if self.check_password()== False:
            self.play_wrong()
            return
        if self.confirm_password() == False:
            self.play_wrong()
            return
        if self.check_city() == False:
            self.play_wrong()
            return
        if self.check_date() == False:
            self.play_wrong()
            return
        if self.checkBoxTerms.isChecked() == False:
            self.play_wrong()
            self.labelException.setText('you have to agree with our TOS')
            return
        else:
            self.show_message_box()

    def show_message_box(self):
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Choose an Option")
        msg_box.setText("what do you want your security question be:")
        
        option1 = msg_box.addButton("city", QMessageBox.ButtonRole.AcceptRole)
        option2 = msg_box.addButton("pet name", QMessageBox.ButtonRole.AcceptRole)
        option3 = msg_box.addButton("favorite food", QMessageBox.ButtonRole.AcceptRole)

        msg_box.exec()

        clicked_button = msg_box.clickedButton()

        if clicked_button == option1:
            self.show_input_dialog("city")
        elif clicked_button == option2:
            self.show_input_dialog("pet name")
        elif clicked_button == option3:
            self.show_input_dialog("favorite food")

    def show_input_dialog(self, option):
        self.option = option
        self.security_answer, ok = QInputDialog.getText(self, f"Input for {option}", f"Enter something for {option}:")
        if ok and self.security_answer.isalpha():
            self.labelException.setText('')
            self.play_correct()
            self.add_memeber()
            self.reset_inputs()
            windowSignUp.close()
            windowLogin.show()

    def reset_inputs(self):
        self.lineFname.setText('')
        self.lineLname.setText('')
        self.linePnumber.setText('')
        self.lineEmail.setText('')
        self.linePass.setText('')
        self.lineConfirmPass.setText('')
        self.lineUsername.setText('')
        self.lineCity.setText('')
        self.labelException.setVisible(False)

    def add_memeber(self):
        fname = self.lineFname.text()
        lname = self.lineLname.text()
        pnumber = self.linePnumber.text()
        email = self.lineEmail.text()
        password = self.linePass.text()
        username = self.lineUsername.text()
        city = self.lineCity.text()
        date = self.date
        security_answer = self.security_answer
        security_type = self.option
        memeber_info = {'first name': [fname], 'last name': [lname], 'phone number':[pnumber], 'username':[username],
        'email':[email], 'password':[password], 'city':[city], 'date':[date], 'security type':[security_type], 'security answer':[security_answer]}

        database_path = project_path + '//database//members_info.xlsx'

        df_new = pd.DataFrame(memeber_info)
    
        # Read existing data
        df_database = pd.read_excel(database_path)
        
        # Append new data
        df_combined = df_database._append(df_new, ignore_index=True)
        
        # Save the combined data to Excel
        df_combined.to_excel(database_path, index=False)

    def check_fname(self):
        fname = self.lineFname.text()
        valid_fname = r'^[a-zA-Z]+$'
        if re.match(valid_fname, fname):
            self.labelException.setText('')
            self.labelFname.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #e0e4e7;
        """)
            return True
        else:
            self.labelException.setText('invalid first name')
            self.labelFname.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #ff0000;
        """)
            return False

    def check_lname(self):
        lname = self.lineLname.text()
        valid_lname = r'^[a-zA-Z]+$'
        if re.match(valid_lname, lname):
            self.labelException.setText('')
            self.labelLname.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #e0e4e7;
        """)
            return True
        else:
            self.labelException.setText('invalid last name')
            self.labelLname.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #ff0000;
            """)
            return False

    def check_pnumber(self):
        pnumber = self.linePnumber.text()
        if pnumber.startswith('09') and pnumber.isnumeric() and len(pnumber) == 11 and pnumber :
            self.labelException.setText('')
            self.labelPnumber.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #e0e4e7;
        """)
            df = pd.read_excel(project_path + '//database//members_info.xlsx')
            if int(pnumber) in df['phone number'].values:
                self.labelException.setText('phone number already in use')
                self.labelPnumber.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #ff0000;
                """)
                return False
            else:
                self.labelException.setText('')
                self.labelPnumber.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #e0e4e7;
                """)
                return True
        else:
            self.labelException.setText('invalid phone number')
            self.labelPnumber.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #ff0000;
            """)
            return False

    def check_email(self):
        email = self.lineEmail.text()
        valid_email = r'^[a-zA-Z0-9._%+-]+@(gmail|yahoo)\.com$'
        if re.match(valid_email, email):
            self.labelException.setText('')
            self.labelEmail.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #e0e4e7;
        """)
            df = pd.read_excel(project_path + '//database//members_info.xlsx')
            if email in df['email'].values:
                self.labelException.setText('email already in use')
                self.labelEmail.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #ff0000;
                """)
                return False
            else:
                self.labelException.setText('')
                self.labelEmail.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #e0e4e7;
                """)
                return True
        else:
            self.labelException.setText('invalid email')
            self.labelEmail.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #ff0000;
            """)
            return False

    #check for changes in password and see if the condition is met:
    def update_labelPassHint(self):
        password = self.linePass.text()
        conditions = {
            'At least one uppercase letter': bool(re.search(r'[A-Z]', password)),
            'At least one lowercase letter': bool(re.search(r'[a-z]', password)),
            'At least one digit': bool(re.search(r'\d', password)),
            'At least one symbol character': bool(re.search(r'[!@#$%^&*()\-_=+{};:,<.>]', password)),
            'At least 8 characters long': len(password) >= 8
        }

        hint_text = ''
        for condition, met in conditions.items():
            color = 'green' if met else 'gray'
            hint_text += f'<span style="color: {color};">{condition}</span><br>'

        self.labelPassHint.setText(hint_text)
        self.labelPassHint.show()

    #clicked on password:
    def on_password_focus_in(self, event):
        self.update_labelPassHint()
        self.labelPassHint.setVisible(True)

    #not clicked:
    def on_password_focus_out(self, event):
        self.labelPassHint.setVisible(False)

    #see/hide password:
    def toggle_echo_mode(self, mode = ''):
        if self.linePass.echoMode() == QLineEdit.EchoMode.Password:
            self.linePass.setEchoMode(QLineEdit.EchoMode.Normal)
            self.buttonToggleEcho.setIcon(QIcon(project_path + "//resources//open eye.ico"))
        else:
            self.linePass.setEchoMode(QLineEdit.EchoMode.Password)
            self.buttonToggleEcho.setIcon(QIcon(project_path + "//resources//close eye.ico"))

        if mode == 'off':
            self.linePass.setEchoMode(QLineEdit.EchoMode.Normal)
            self.buttonToggleEcho.setIcon(QIcon(project_path + "//resources//open eye.ico"))
        elif mode == 'on':
            self.linePass.setEchoMode(QLineEdit.EchoMode.Password)
            self.buttonToggleEcho.setIcon(QIcon(project_path + "//resources//close eye.ico"))
        self.on_password_focus_in(self.event)

    def check_password(self):
        valid_password = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$'
        password = self.linePass.text()
        if re.match(valid_password, password):
            self.labelException.setText('')
            self.labelPass.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #e0e4e7;
        """)
            return True
        else:
            self.labelException.setText('invalid password')
            self.labelPass.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #ff0000;
            """)
            return False
    
    def check_username(self):
        username = self.lineUsername.text()
        valid_username = r'^(?=.*[a-zA-Z])[a-zA-Z0-9-]+$'
        if re.match(valid_username, username):
            self.labelException.setText('')
            self.labelUsername.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #e0e4e7;
        """)
            df = pd.read_excel(project_path + '//database//members_info.xlsx')
            usernames_inlower = df['username'].str.lower().tolist()
            if username.lower() in usernames_inlower:
                self.labelException.setText('username taken')
                self.labelUsername.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #ff0000;
                """)
                return False
            else:
                self.labelException.setText('')
                self.labelUsername.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #e0e4e7;
                """)
                return True
        else:
            self.labelException.setText("invalid username")
            self.labelUsername.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #ff0000;
            """)
            return False

    def confirm_password(self):
        password = self.linePass.text()
        retyped_password = self.lineConfirmPass.text()
        if retyped_password == password:
            self.labelException.setText('')
            self.labelConfirmPass.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #e0e4e7;
        """)
            return True
        else: 
            self.labelException.setText("Retyped password doesn't match the original one")
            self.labelConfirmPass.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #ff0000;
            """)
            return False
                     
    def check_city(self):
        city = self.lineCity.text()
        if city in self.city_list:
            self.labelException.setText('')
            self.labelCity.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #e0e4e7;
            """)
            return True
        else:
            self.labelException.setText('invalid date')
            self.labelCity.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #ff0000;
            """)
            return False

    def check_date(self):
        day = int(self.comboDay.currentText())
        month = self.comboMonth.currentText()
        year = int(self.comboYear.currentText())
        if 0 < day <= self.return_max_day(month, year):
            self.labelException.setText('')
            self.labelDate.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #e0e4e7;
        """)
            self.date = str(year) + '/' + month + '/' + str(day)
            return True
        else:
            self.labelException.setText('invalid date')
            self.labelDate.setStyleSheet("""
            background-color:rgba( 255, 255, 255, 10% );
            border-radius: 8px;
            padding: 5px 15px;
            border: 1px solid #ff0000;
            """)
            return False

    def isLeapYear(self, year):
        leapList = list()
        for i in range(1920, 2006, 4):
            leapList.append(i)
        if year in leapList:
            return True
        else:
            return False

    def return_max_day(self, month, year):
        month_dict = {
            'January' : 31,
            'February' : 28,
            'March' : 31,
            'April' : 30,
            'May' : 31,
            'June' : 30,
            'July' : 31,
            'August' : 31,
            'September' : 30,
            'October' : 31,
            'November' : 30,
            'December' : 31,
        }
        if self.isLeapYear(year) and month == 'February':
            return 29
        else:
            return month_dict[month]
    

#Login ui:
class LoginPage(QWidget):
    def __init__(self):
        super().__init__()

        #ui/tile/icon setup:
        uic.loadUi(project_path + "//LoginPage.ui", self)
        self.setWindowIcon(QIcon(project_path + "//resources//login icon.ico"))
        self.buttonMute.setIcon(QIcon(project_path + "//resources//sound.ico"))
        self.setWindowTitle('Login')

        #username:
        self.username = ''

        #background image setup:
        self.labelPic = QLabel(self)
        self.labelPic.resize(16777215, 16777215)
        self.labelPic.setStyleSheet(star_theme_path)
        self.labelPic.lower()

        #exception handling label:
        self.labelException.setVisible(False)
        self.attempts = 0

        #sound initilization:
        self.wrong_sound = Sound('Alert')
        self.wrong_sound_isMuted = False
        self.correct_sound = Sound('Correct')
        self.correct_sound_isMuted = False

        #signal handling:
        self.labelPassForgot.mousePressEvent = self.open_passForgot
        self.buttonLogin.clicked.connect(self.check_login_input)
        self.buttonSignUp.clicked.connect(self.open_signUp_page)
        self.buttonMute.clicked.connect(self.play_mute_background)
        self.buttonLogin.setShortcut("Return")
        self.buttonToggleEcho.clicked.connect(self.toggle_echo_mode)
        self.buttonToggleEcho.setIcon(QIcon(project_path + "//resources//close eye.ico"))

    def reset_inputs(self):
        self.linePassword.setText('')
        self.lineUsername.setText('')

    def play_mute_background(self):
        if backgroundSound.IsMuted():
            backgroundSound.Mute(False)
            self.buttonMute.setIcon(QIcon(project_path + "//resources//sound.ico"))
            windowSignUp.buttonMute.setIcon(QIcon(project_path + "//resources//sound.ico"))
            windowMain.buttonMute.setIcon(QIcon(project_path + "//resources//sound.ico"))
            self.wrong_sound_isMuted = False
            self.correct_sound_isMuted = False
            SignUp.wrong_sound_isMuted = False
        else:
            backgroundSound.Mute(True)
            self.buttonMute.setIcon(QIcon(project_path + "//resources//mute sound.ico"))
            windowSignUp.buttonMute.setIcon(QIcon(project_path + "//resources//mute sound.ico"))
            windowMain.buttonMute.setIcon(QIcon(project_path + "//resources//mute sound.ico"))
            self.wrong_sound_isMuted = True
            self.correct_sound_isMuted == True
            SignUp.wrong_sound_isMuted = True

    def toggle_echo_mode(self):
        if self.linePassword.echoMode() == QLineEdit.EchoMode.Password:
            self.linePassword.setEchoMode(QLineEdit.EchoMode.Normal)
            self.buttonToggleEcho.setIcon(QIcon(project_path + "//resources//open eye.ico"))
        else:
            self.linePassword.setEchoMode(QLineEdit.EchoMode.Password)
            self.buttonToggleEcho.setIcon(QIcon(project_path + "//resources//close eye.ico"))

    def play_wrong(self):
        if self.wrong_sound_isMuted == False:
            self.wrong_sound.Play()
    
    def play_correct(self):
        if self.correct_sound_isMuted == False:
            self.correct_sound.Play()

    def open_passForgot(self,*arg, **kwargs):
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Recovery")
        msg_box.setText("choose a recovery option:")
        
        option1 = msg_box.addButton("send message", QMessageBox.ButtonRole.AcceptRole)
        option2 = msg_box.addButton("ask a question", QMessageBox.ButtonRole.AcceptRole)
        option3 = msg_box.addButton("cancel", QMessageBox.ButtonRole.AcceptRole)
        msg_box.exec()

        clicked_button = msg_box.clickedButton()

        if clicked_button == option1:
            windowLogin.close()
            windowPassRecovery.show()
        elif clicked_button == option2:
            self.show_message_recovery()
        elif clicked_button == option3:
            return

    def open_signUp_page(self):
        windowSignUp.show()
        windowLogin.close()
    
    def show_message_recovery(self):
        username, ok = QInputDialog.getText(self, 'Username Input', 'Enter username:')
        if not ok:
            return

        if not username:
            QMessageBox.warning(self, 'Input Error', 'Please enter a username.')
            return

        try:
            df = pd.read_excel(project_path + '//database//members_info.xlsx')
        except Exception as e:
            QMessageBox.critical(self, 'File Error', f'Error loading file: {e}')
            return

        user_row = df[df['username'] == username]
        if user_row.empty:
            QMessageBox.warning(self, 'Input Error', 'username not found.')
        else:
            self.username = username
            self.security_type = user_row.iloc[0]['security type']
            self.security_answer = user_row.iloc[0]['security answer']
            self.check_security_answer()
    
    def check_security_answer(self):
        answer, ok = QInputDialog.getText(self, 'Answer Input', 'Enter your answer for ' + str(self.security_type) + ':')
        if not ok:
            return
        while(answer != self.security_answer):
            QMessageBox.warning(self, 'Input Error', 'Invalid answer')
            answer, ok = QInputDialog.getText(self, 'Answer Input', 'Enter your answer for ' + str(self.security_type) + ':')
            if not ok:
                break
        else:
            self.play_correct()
            windowLogin.close()
            self.reset_inputs()
            windowMain.reinit()
            windowMain.show()
        return

    #function to check inputs:
    def check_login_input(self):
        if self.attempts < 3:
            self.username = self.lineUsername.text()
            self.password = self.linePassword.text()
            valid_password = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$'
            if 0 < len(self.username):
                self.labelException.setVisible(False)
                self.labelException.setText('')
                self.labelUsername.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #e0e4e7;
                """)
            else:
                self.labelException.setVisible(True)
                self.labelException.setText('invalid username')
                self.labelUsername.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #ff0000;
                """)
                self.play_wrong()
                self.lock()
                return

            if re.match(valid_password, self.password):
                self.labelException.setVisible(False)
                self.labelException.setText('')
                self.labelPassword.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #e0e4e7;
                """)

                if self.check_login(self.username, self.password):
                    self.play_correct()
                    windowLogin.close()
                    self.reset_inputs()
                    windowMain.reinit()
                    windowMain.show()
                else:
                    self.lock()
                    self.play_wrong()
                    return
            else:
                self.labelException.setVisible(True)
                self.labelException.setText('invalid password')
                self.labelPassword.setStyleSheet("""
                background-color:rgba( 255, 255, 255, 10% );
                border-radius: 8px;
                padding: 5px 15px;
                border: 1px solid #ff0000;
                """)
                self.play_wrong()
                self.lock()
                return
                
        else:
            self.countdown()

    def check_login(self, username, password):
        try:
            # Read the Excel file
            df = pd.read_excel(project_path + '//database//members_info.xlsx')
            
            # Check if the DataFrame has the required columns
            if 'username' not in df.columns or 'password' not in df.columns:
                self.labelException.setVisible(True)
                self.labelException.setText('Corrupted database')
                return False
       
            # Check if there is a matching row
            user_row = df[(df['username'] == username) & (df['password'] == password)]
            if not user_row.empty:
                self.labelException.setText('')
                return True
            else:
                self.labelException.setVisible(True)
                if username in df['username'].values():
                    self.labelException.setText('password is incorrect (forgot your password ?)')
                else:
                    self.labelException.setText('username not found (you may want to sign up)')
                return False
        except Exception as e:
            print(e)
            return False
            
    def lock(self):
        print(self.attempts)
        self.attempts += 1
        if self.attempts == 3:
            self.labelException.setText("max attemts reached!")
            self.stored_time1 = time.time()
            self.countdown()

    def countdown(self):
        self.stored_time2 = time.time()
        self.time_difference = self.stored_time2 - self.stored_time1
        if int(self.time_difference) < 60:
            self.labelException.setVisible(True)
            self.labelException.setText('please try again in: ' + str(int(60 - self.time_difference)) + ' second(s)') 
        else:
            self.attempts = 0
            self.labelException.setText('')
            self.check_login_input()
            
class PassRecovery(QWidget):
    def __init__(self):
        super().__init__()

        #ui/tile/icon setup:
        uic.loadUi(project_path + "//PasswordRecoveryPage.ui", self)
        self.setWindowTitle('Recovery')
        self.setWindowIcon(QIcon(project_path + "//resources//login icon.ico"))
        
        #background image setup:
        self.labelPic = QLabel(self)
        self.labelPic.resize(16777215, 16777215)
        self.labelPic.setStyleSheet(star_theme_path)
        self.labelPic.lower()

        #signal handling:
        self.labelException.setVisible(False)
        self.buttonVia.clicked.connect(self.send_via)
        self.buttonBack.clicked.connect(self.go_back)
        self.buttonSend.clicked.connect(self.check)

    def send_via(self):
        self.labelException.setVisible(False)
        self.lineInput.setText('')
        if self.labelInput.text() == 'Email:':
            self.labelInput.setText('Phone number:')
            self.buttonVia.setText('Send via Email')
        else:
            self.labelInput.setText('Email:')
            self.buttonVia.setText('Send via SMS')
            
    def check(self):
        self.labelException.setVisible(True)
        if self.labelInput.text() == 'Email:':
            if self.check_email():
                self.labelException.setText('we have sent a message to your email address seccessfully')
            else:
                self.labelException.setText('email not found')
        else:
            if self.check_pnumber():
                self.labelException.setText('we have sent a SMS to your number seccessfully')
            else:
                self.labelException.setText('phone number not found')

    def check_pnumber(self):
        pnumber = self.lineInput.text()
        if pnumber.isnumeric() == False:
            return False
        else:
            df = pd.read_excel(project_path + '//database//members_info.xlsx')
            if int(pnumber) in df['phone number'].values:
                return True
            else:
                return False
    
    def check_email(self):
        email = self.lineInput.text()
        df = pd.read_excel(project_path + '//database//members_info.xlsx')
        if email in df['email'].values:
            return True
        else:
            return False

    def go_back(self):
        windowPassRecovery.close()
        windowLogin.show()

#main:
if __name__ == '__main__':
    app = QApplication(sys.argv)

    #background music initialization:
    backgroundSound = Sound('background music')
    backgroundSound.SetLoop(-2)
    backgroundSound.Play()

    #define windows:
    windowLogin = LoginPage()
    windowMain = MainApp()
    windowSignUp = SignUp()
    windowPassRecovery = PassRecovery()
    
    #show window(s):
    windowLogin.show()
    #windowMain.show()

    #exit:
    try:
        sys.exit(app.exec())
    except SystemExit:
        print('Closing Window...')